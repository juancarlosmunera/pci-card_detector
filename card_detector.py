"""
PCI-DSS Cardholder Data Discovery Tool
Purpose: Detect potential credit card numbers in datasets for compliance and risk management
Uses Luhn algorithm (mod-10) for validation

Supported datasources:
  Files     : CSV, TXT, LOG, JSON, XML, SQL, PDF, Excel
  Databases : SQLite, PostgreSQL, MySQL
  Cloud     : Amazon S3, Google Cloud Storage, Azure Blob Storage

Optional dependencies (install only what you need):
  pip install psycopg2-binary          # PostgreSQL
  pip install mysql-connector-python   # MySQL
  pip install pdfplumber               # PDF files
  pip install openpyxl                 # Excel files
  pip install boto3                    # Amazon S3
  pip install google-cloud-storage     # Google Cloud Storage
  pip install azure-storage-blob       # Azure Blob Storage
"""

import re
import csv
import os
import sqlite3
import tempfile
from pathlib import Path
from typing import List, Dict
import argparse

# ── Optional dependency detection ────────────────────────────────────────────

try:
    import psycopg2
    HAS_POSTGRES = True
except ImportError:
    HAS_POSTGRES = False

try:
    import mysql.connector
    HAS_MYSQL = True
except ImportError:
    HAS_MYSQL = False

try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    import boto3
    HAS_S3 = True
except ImportError:
    HAS_S3 = False

try:
    from google.cloud import storage as gcs_storage
    HAS_GCS = True
except ImportError:
    HAS_GCS = False

try:
    from azure.storage.blob import BlobServiceClient
    HAS_AZURE = True
except ImportError:
    HAS_AZURE = False

# ── Scannable extensions (used by directory and cloud scanners) ───────────────

SCANNABLE_EXTENSIONS = {'.csv', '.txt', '.log', '.json', '.xml', '.sql', '.pdf', '.xlsx'}


class CreditCardDetector:
    """
    Detects and validates credit card numbers using the Luhn algorithm.
    Supports modern BIN formats (6-8 digit BINs).

    Datasources: local files (CSV/TXT/LOG/JSON/XML/SQL/PDF/Excel),
                 databases (SQLite/PostgreSQL/MySQL), and
                 cloud storage (S3/GCS/Azure Blob).
    """

    # Card brand patterns (IIN ranges)
    CARD_PATTERNS = {
        'Visa': r'^4[0-9]{12}(?:[0-9]{3})?$',                          # 13 or 16 digits
        'Mastercard': r'^5[1-5][0-9]{14}$|^2[2-7][0-9]{14}$',         # 16 digits
        'Amex': r'^3[47][0-9]{13}$',                                    # 15 digits
        'Discover': r'^6(?:011|5[0-9]{2})[0-9]{12}$',                  # 16 digits
        'Diners': r'^3(?:0[0-5]|[68][0-9])[0-9]{11}$',                 # 14 digits
        'JCB': r'^(?:2131|1800|35\d{3})\d{11}$',                       # 16 digits
    }

    def __init__(self):
        self.findings = []

    # ── Core detection ────────────────────────────────────────────────────────

    def luhn_check(self, card_number: str) -> bool:
        """
        Validates a card number using the Luhn algorithm (mod-10).

        The Luhn algorithm:
        1. Starting from the rightmost digit (check digit), move left
        2. Double every second digit
        3. If doubling results in a number > 9, subtract 9
        4. Sum all digits
        5. If sum % 10 == 0, the number is valid

        Args:
            card_number: String of digits (no spaces/dashes)

        Returns:
            bool: True if valid per Luhn algorithm
        """
        card_number = re.sub(r'[\s-]', '', card_number)

        if not card_number.isdigit():
            return False

        if len(card_number) < 13 or len(card_number) > 19:
            return False

        digits = [int(d) for d in card_number]
        checksum = 0

        for i in range(len(digits) - 1, -1, -1):
            digit = digits[i]
            if (len(digits) - i) % 2 == 0:
                digit *= 2
                if digit > 9:
                    digit -= 9
            checksum += digit

        return checksum % 10 == 0

    def identify_card_brand(self, card_number: str) -> str:
        """
        Identifies the card brand based on IIN (Issuer Identification Number).

        Args:
            card_number: Validated card number

        Returns:
            str: Card brand name or 'Unknown'
        """
        for brand, pattern in self.CARD_PATTERNS.items():
            if re.match(pattern, card_number):
                return brand
        return 'Unknown'

    def find_card_numbers(self, text: str) -> List[Dict]:
        """
        Searches text for potential card numbers and validates them.

        Args:
            text: Text to search

        Returns:
            List of dicts containing found card numbers and metadata
        """
        findings = []
        pattern = r'\b\d{4}[\s-]?\d{4}[\s-]?\d{4}[\s-]?\d{4,7}\b'

        for match in re.finditer(pattern, text):
            candidate = match.group()
            clean_number = re.sub(r'[\s-]', '', candidate)

            if self.luhn_check(clean_number):
                brand = self.identify_card_brand(clean_number)
                masked = f"{clean_number[:6]}...{clean_number[-4:]}"

                findings.append({
                    'original_format': candidate,
                    'masked_number': masked,
                    'card_brand': brand,
                    'position': match.start(),
                    'length': len(clean_number),
                })

        return findings

    # ── File scanners ─────────────────────────────────────────────────────────

    def scan_csv(self, csv_path: str, delimiter: str = ',') -> List[Dict]:
        """
        Scans a CSV file for credit card numbers.

        Args:
            csv_path: Path to CSV file
            delimiter: CSV delimiter (default: comma)

        Returns:
            List of findings with location information
        """
        findings = []
        try:
            with open(csv_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f, delimiter=delimiter)
                for row_num, row in enumerate(reader, start=1):
                    for col_num, cell in enumerate(row, start=1):
                        for finding in self.find_card_numbers(str(cell)):
                            finding['file'] = csv_path
                            finding['row'] = row_num
                            finding['column'] = col_num
                            finding['cell_content'] = cell[:50]
                            findings.append(finding)
        except Exception as e:
            print(f"Error scanning {csv_path}: {e}")
        return findings

    def scan_text_file(self, file_path: str) -> List[Dict]:
        """
        Scans a text file for credit card numbers.
        Supports .txt, .log, .json, .xml, .sql, etc.

        Args:
            file_path: Path to text file

        Returns:
            List of findings with location information
        """
        findings = []
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line_num, line in enumerate(f, start=1):
                    for finding in self.find_card_numbers(line):
                        finding['file'] = file_path
                        finding['line'] = line_num
                        finding['context'] = line.strip()[:100]
                        findings.append(finding)
        except Exception as e:
            print(f"Error scanning {file_path}: {e}")
        return findings

    def scan_pdf(self, pdf_path: str) -> List[Dict]:
        """
        Scans a PDF file for credit card numbers by extracting text page by page.

        Requires: pip install pdfplumber

        Args:
            pdf_path: Path to PDF file

        Returns:
            List of findings with page and line location
        """
        if not HAS_PDF:
            print("PDF support requires: pip install pdfplumber")
            return []

        findings = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, start=1):
                    text = page.extract_text()
                    if not text:
                        continue
                    for line_num, line in enumerate(text.split('\n'), start=1):
                        for finding in self.find_card_numbers(line):
                            finding['file'] = pdf_path
                            finding['page'] = page_num
                            finding['line'] = line_num
                            finding['context'] = line.strip()[:100]
                            findings.append(finding)
        except Exception as e:
            print(f"Error scanning {pdf_path}: {e}")
        return findings

    def scan_excel(self, excel_path: str) -> List[Dict]:
        """
        Scans an Excel workbook (.xlsx) for credit card numbers across all sheets.

        Requires: pip install openpyxl

        Args:
            excel_path: Path to Excel file

        Returns:
            List of findings with sheet, row, and column location
        """
        if not HAS_EXCEL:
            print("Excel support requires: pip install openpyxl")
            return []

        findings = []
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row_num, row in enumerate(sheet.iter_rows(), start=1):
                    for col_num, cell in enumerate(row, start=1):
                        if cell.value is None:
                            continue
                        cell_str = str(cell.value)
                        for finding in self.find_card_numbers(cell_str):
                            finding['file'] = excel_path
                            finding['sheet'] = sheet_name
                            finding['row'] = row_num
                            finding['column'] = col_num
                            finding['cell_content'] = cell_str[:50]
                            findings.append(finding)
            wb.close()
        except Exception as e:
            print(f"Error scanning {excel_path}: {e}")
        return findings

    def _scan_file_by_extension(self, local_path: str, ext: str,
                                 source: str = None) -> List[Dict]:
        """
        Routes a local file to the correct scanner based on its extension.
        Optionally overrides the 'file' field with a source URI (e.g. s3://...).
        """
        if ext == '.csv':
            findings = self.scan_csv(local_path)
        elif ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
            findings = self.scan_excel(local_path)
        elif ext == '.pdf':
            findings = self.scan_pdf(local_path)
        else:
            findings = self.scan_text_file(local_path)

        if source:
            for f in findings:
                f['file'] = source
        return findings

    def scan_directory(self, directory: str, extensions: List[str] = None) -> List[Dict]:
        """
        Recursively scans a directory for files containing card numbers.

        Args:
            directory: Directory path to scan
            extensions: File extensions to scan. Defaults to all supported types
                        including PDF and Excel.

        Returns:
            List of all findings
        """
        if extensions is None:
            extensions = list(SCANNABLE_EXTENSIONS)

        all_findings = []
        for file_path in Path(directory).rglob('*'):
            if file_path.is_file() and file_path.suffix.lower() in extensions:
                print(f"Scanning: {file_path}")
                all_findings.extend(
                    self._scan_file_by_extension(str(file_path), file_path.suffix.lower())
                )
        return all_findings

    # ── Database scanners ─────────────────────────────────────────────────────

    def scan_sqlite(self, db_path: str, row_limit: int = 10000) -> List[Dict]:
        """
        Scans all tables in a SQLite database for credit card numbers.

        SQLite is part of the Python standard library — no extra install needed.

        Args:
            db_path:   Path to the SQLite database file
            row_limit: Maximum rows to scan per table (default: 10 000)

        Returns:
            List of findings with table, column, and row ID location
        """
        findings = []
        source = f"sqlite:{db_path}"
        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [r[0] for r in cur.fetchall()]

            for table in tables:
                cur.execute(f'PRAGMA table_info("{table}")')
                columns = [row[1] for row in cur.fetchall()]
                if not columns:
                    continue

                try:
                    cur.execute(f'SELECT rowid, * FROM "{table}" LIMIT {row_limit}')
                    for row in cur.fetchall():
                        rowid = row[0]
                        for col_idx, cell_value in enumerate(row[1:]):
                            if cell_value is None:
                                continue
                            for finding in self.find_card_numbers(str(cell_value)):
                                finding['source'] = source
                                finding['table'] = table
                                finding['column'] = columns[col_idx]
                                finding['row_id'] = rowid
                                findings.append(finding)
                except Exception as e:
                    print(f"  Error scanning table '{table}': {e}")

            conn.close()
        except Exception as e:
            print(f"Error opening SQLite database {db_path}: {e}")
        return findings

    def scan_postgres(self, host: str, dbname: str, user: str, password: str,
                      port: int = 5432, schema: str = 'public',
                      row_limit: int = 10000) -> List[Dict]:
        """
        Scans text columns in a PostgreSQL database for credit card numbers.

        Requires: pip install psycopg2-binary

        Args:
            host, dbname, user, password: Connection credentials
            port:      PostgreSQL port (default: 5432)
            schema:    Schema to scan (default: public)
            row_limit: Maximum rows to scan per table (default: 10 000)

        Returns:
            List of findings with table, column, and ctid location
        """
        if not HAS_POSTGRES:
            print("PostgreSQL support requires: pip install psycopg2-binary")
            return []

        findings = []
        source = f"postgres:{host}/{dbname}"
        try:
            conn = psycopg2.connect(host=host, port=port, dbname=dbname,
                                    user=user, password=password)
            cur = conn.cursor()

            cur.execute("""
                SELECT table_name, column_name
                FROM information_schema.columns
                WHERE table_schema = %s
                  AND data_type IN ('character varying', 'text', 'character',
                                    'varchar', 'name')
                ORDER BY table_name, column_name
            """, (schema,))

            table_columns: Dict[str, List[str]] = {}
            for table, column in cur.fetchall():
                table_columns.setdefault(table, []).append(column)

            for table, columns in table_columns.items():
                col_list = ', '.join(f'"{c}"' for c in columns)
                try:
                    cur.execute(
                        f'SELECT ctid, {col_list} FROM "{schema}"."{table}" LIMIT {row_limit}'
                    )
                    for row in cur.fetchall():
                        ctid = row[0]
                        for col_idx, cell_value in enumerate(row[1:]):
                            if cell_value is None:
                                continue
                            for finding in self.find_card_numbers(str(cell_value)):
                                finding['source'] = source
                                finding['table'] = f"{schema}.{table}"
                                finding['column'] = columns[col_idx]
                                finding['row_id'] = str(ctid)
                                findings.append(finding)
                except Exception as e:
                    print(f"  Error scanning {schema}.{table}: {e}")

            conn.close()
        except Exception as e:
            print(f"Error connecting to PostgreSQL {host}/{dbname}: {e}")
        return findings

    def scan_mysql(self, host: str, database: str, user: str, password: str,
                   port: int = 3306, row_limit: int = 10000) -> List[Dict]:
        """
        Scans text columns in a MySQL database for credit card numbers.

        Requires: pip install mysql-connector-python

        Args:
            host, database, user, password: Connection credentials
            port:      MySQL port (default: 3306)
            row_limit: Maximum rows to scan per table (default: 10 000)

        Returns:
            List of findings with table, column, and approximate row location
        """
        if not HAS_MYSQL:
            print("MySQL support requires: pip install mysql-connector-python")
            return []

        findings = []
        source = f"mysql:{host}/{database}"
        try:
            conn = mysql.connector.connect(
                host=host, port=port, database=database,
                user=user, password=password
            )
            cur = conn.cursor()

            cur.execute("""
                SELECT TABLE_NAME, COLUMN_NAME
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = %s
                  AND DATA_TYPE IN ('varchar', 'text', 'char', 'longtext',
                                    'mediumtext', 'tinytext')
                ORDER BY TABLE_NAME, COLUMN_NAME
            """, (database,))

            table_columns: Dict[str, List[str]] = {}
            for table, column in cur.fetchall():
                table_columns.setdefault(table, []).append(column)

            for table, columns in table_columns.items():
                col_list = ', '.join(f'`{c}`' for c in columns)
                try:
                    cur.execute(f'SELECT {col_list} FROM `{table}` LIMIT {row_limit}')
                    for row_num, row in enumerate(cur.fetchall(), start=1):
                        for col_idx, cell_value in enumerate(row):
                            if cell_value is None:
                                continue
                            for finding in self.find_card_numbers(str(cell_value)):
                                finding['source'] = source
                                finding['table'] = table
                                finding['column'] = columns[col_idx]
                                finding['row_id'] = row_num
                                findings.append(finding)
                except Exception as e:
                    print(f"  Error scanning {table}: {e}")

            conn.close()
        except Exception as e:
            print(f"Error connecting to MySQL {host}/{database}: {e}")
        return findings

    # ── Cloud storage scanners ────────────────────────────────────────────────

    def scan_s3(self, bucket: str, prefix: str = '',
                region: str = None, aws_access_key_id: str = None,
                aws_secret_access_key: str = None) -> List[Dict]:
        """
        Scans files in an Amazon S3 bucket for credit card numbers.

        Requires: pip install boto3
        Credentials are read from the standard AWS credential chain
        (env vars, ~/.aws/credentials, IAM role) unless overridden.

        Args:
            bucket:               S3 bucket name
            prefix:               Key prefix to filter objects (default: all)
            region:               AWS region override
            aws_access_key_id:    Optional explicit AWS key
            aws_secret_access_key: Optional explicit AWS secret

        Returns:
            List of findings; 'file' field contains the s3:// URI
        """
        if not HAS_S3:
            print("S3 support requires: pip install boto3")
            return []

        client_kwargs = {}
        if region:
            client_kwargs['region_name'] = region
        if aws_access_key_id:
            client_kwargs['aws_access_key_id'] = aws_access_key_id
        if aws_secret_access_key:
            client_kwargs['aws_secret_access_key'] = aws_secret_access_key

        s3 = boto3.client('s3', **client_kwargs)
        findings = []

        try:
            paginator = s3.get_paginator('list_objects_v2')
            with tempfile.TemporaryDirectory() as tmpdir:
                for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
                    for obj in page.get('Contents', []):
                        key = obj['Key']
                        ext = Path(key).suffix.lower()
                        if ext not in SCANNABLE_EXTENSIONS:
                            continue

                        source = f"s3://{bucket}/{key}"
                        local_path = os.path.join(tmpdir, Path(key).name)
                        print(f"  Downloading: {source}")
                        try:
                            s3.download_file(bucket, key, local_path)
                            findings.extend(
                                self._scan_file_by_extension(local_path, ext, source)
                            )
                        except Exception as e:
                            print(f"  Error processing {key}: {e}")
        except Exception as e:
            print(f"Error accessing S3 bucket '{bucket}': {e}")
        return findings

    def scan_gcs(self, bucket: str, prefix: str = '') -> List[Dict]:
        """
        Scans files in a Google Cloud Storage bucket for credit card numbers.

        Requires: pip install google-cloud-storage
        Credentials are read from Application Default Credentials
        (GOOGLE_APPLICATION_CREDENTIALS env var or gcloud auth).

        Args:
            bucket: GCS bucket name
            prefix: Blob name prefix to filter (default: all)

        Returns:
            List of findings; 'file' field contains the gs:// URI
        """
        if not HAS_GCS:
            print("GCS support requires: pip install google-cloud-storage")
            return []

        findings = []
        try:
            client = gcs_storage.Client()
            with tempfile.TemporaryDirectory() as tmpdir:
                for blob in client.list_blobs(bucket, prefix=prefix):
                    ext = Path(blob.name).suffix.lower()
                    if ext not in SCANNABLE_EXTENSIONS:
                        continue

                    source = f"gs://{bucket}/{blob.name}"
                    local_path = os.path.join(tmpdir, Path(blob.name).name)
                    print(f"  Downloading: {source}")
                    try:
                        blob.download_to_filename(local_path)
                        findings.extend(
                            self._scan_file_by_extension(local_path, ext, source)
                        )
                    except Exception as e:
                        print(f"  Error processing {blob.name}: {e}")
        except Exception as e:
            print(f"Error accessing GCS bucket '{bucket}': {e}")
        return findings

    def scan_azure_blob(self, container: str, prefix: str = '',
                        connection_string: str = None) -> List[Dict]:
        """
        Scans files in an Azure Blob Storage container for credit card numbers.

        Requires: pip install azure-storage-blob
        Connection string is read from --azure-conn-string or the
        AZURE_STORAGE_CONNECTION_STRING environment variable.

        Args:
            container:         Azure Blob container name
            prefix:            Blob name prefix to filter (default: all)
            connection_string: Azure Storage connection string

        Returns:
            List of findings; 'file' field contains the azure:// URI
        """
        if not HAS_AZURE:
            print("Azure Blob support requires: pip install azure-storage-blob")
            return []

        if not connection_string:
            connection_string = os.environ.get('AZURE_STORAGE_CONNECTION_STRING')
        if not connection_string:
            print("Azure connection string required via --azure-conn-string "
                  "or AZURE_STORAGE_CONNECTION_STRING environment variable.")
            return []

        findings = []
        try:
            service_client = BlobServiceClient.from_connection_string(connection_string)
            container_client = service_client.get_container_client(container)

            with tempfile.TemporaryDirectory() as tmpdir:
                for blob in container_client.list_blobs(name_starts_with=prefix or None):
                    ext = Path(blob.name).suffix.lower()
                    if ext not in SCANNABLE_EXTENSIONS:
                        continue

                    source = f"azure://{container}/{blob.name}"
                    local_path = os.path.join(tmpdir, Path(blob.name).name)
                    print(f"  Downloading: {source}")
                    try:
                        blob_client = container_client.get_blob_client(blob.name)
                        with open(local_path, 'wb') as f:
                            f.write(blob_client.download_blob().readall())
                        findings.extend(
                            self._scan_file_by_extension(local_path, ext, source)
                        )
                    except Exception as e:
                        print(f"  Error processing {blob.name}: {e}")
        except Exception as e:
            print(f"Error accessing Azure container '{container}': {e}")
        return findings

    # ── Reporting ─────────────────────────────────────────────────────────────

    def generate_report(self, findings: List[Dict], output_path: str = None):
        """
        Prints a formatted report of findings to stdout and optionally saves a CSV.

        Args:
            findings:    List of detection findings
            output_path: Optional path to save a CSV report
        """
        if not findings:
            print("\n[OK] No credit card numbers detected.")
            return

        print(f"\n[WARNING] {len(findings)} potential credit card number(s) detected!\n")
        print("=" * 80)

        for i, finding in enumerate(findings, start=1):
            print(f"\nFinding #{i}:")

            # Source label
            source = finding.get('source') or finding.get('file', 'N/A')
            print(f"  Source   : {source}")

            # Location — varies by datasource
            if 'table' in finding:
                print(f"  Location : Table={finding['table']}, "
                      f"Column={finding['column']}, RowID={finding.get('row_id', 'N/A')}")
            elif 'sheet' in finding:
                print(f"  Location : Sheet={finding['sheet']}, "
                      f"Row={finding['row']}, Column={finding['column']}")
            elif 'page' in finding:
                print(f"  Location : Page={finding['page']}, "
                      f"Line={finding.get('line', 'N/A')}")
            elif 'row' in finding:
                print(f"  Location : Row={finding['row']}, Column={finding['column']}")
            elif 'line' in finding:
                print(f"  Location : Line={finding['line']}")

            print(f"  Masked   : {finding['masked_number']}")
            print(f"  Brand    : {finding['card_brand']}")
            print(f"  Format   : {finding['original_format']}")
            print(f"  Length   : {finding['length']} digits")

        print("\n" + "=" * 80)

        if output_path:
            self.save_report_csv(findings, output_path)
            print(f"\n[OK] Report saved to: {output_path}")

    def save_report_csv(self, findings: List[Dict], output_path: str):
        """
        Saves findings to a CSV report.

        Args:
            findings:    List of detection findings
            output_path: Path for output CSV file
        """
        if not findings:
            return

        fieldnames = [
            'source', 'file', 'table', 'column', 'row_id',
            'sheet', 'row', 'line', 'page',
            'masked_number', 'card_brand', 'original_format', 'length',
            'context', 'cell_content',
        ]

        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(findings)


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    """
    Main CLI interface for the card detection tool.
    """
    parser = argparse.ArgumentParser(
        description='PCI-DSS Cardholder Data Discovery Tool',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # File sources
  python card_detector.py --csv transactions.csv
  python card_detector.py --csv data.tsv --delimiter $'\\t'
  python card_detector.py --file application.log
  python card_detector.py --pdf invoice.pdf
  python card_detector.py --excel report.xlsx
  python card_detector.py --directory /path/to/data --output findings.csv

  # Databases
  python card_detector.py --sqlite /var/db/app.db
  python card_detector.py --pg-host localhost --pg-db mydb --pg-user alice --pg-password s3cr3t
  python card_detector.py --mysql-host 10.0.0.5 --mysql-db sales --mysql-user root --mysql-password pass

  # Cloud storage (credentials via env/config -- see docs)
  python card_detector.py --s3-bucket my-bucket --s3-prefix exports/
  python card_detector.py --gcs-bucket my-bucket
  python card_detector.py --azure-container backups --azure-conn-string "DefaultEndpoints..."
"""
    )

    # ── File sources ──────────────────────────────────────────────────────────
    file_group = parser.add_argument_group('File sources')
    file_group.add_argument('--csv', metavar='FILE', help='Scan a CSV file')
    file_group.add_argument('--file', metavar='FILE',
                            help='Scan a text/log/JSON/XML/SQL file')
    file_group.add_argument('--pdf', metavar='FILE',
                            help='Scan a PDF file (requires pdfplumber)')
    file_group.add_argument('--excel', metavar='FILE',
                            help='Scan an Excel workbook (.xlsx) (requires openpyxl)')
    file_group.add_argument('--directory', metavar='PATH',
                            help='Recursively scan all supported files in a directory')
    file_group.add_argument('--delimiter', default=',',
                            help='CSV delimiter (default: comma)')

    # ── Database sources ──────────────────────────────────────────────────────
    db_group = parser.add_argument_group('Database sources')
    db_group.add_argument('--sqlite', metavar='FILE',
                          help='Scan a SQLite database file')

    db_group.add_argument('--pg-host', metavar='HOST', help='PostgreSQL host')
    db_group.add_argument('--pg-port', metavar='PORT', type=int, default=5432,
                          help='PostgreSQL port (default: 5432)')
    db_group.add_argument('--pg-db', metavar='DBNAME', help='PostgreSQL database name')
    db_group.add_argument('--pg-user', metavar='USER', help='PostgreSQL user')
    db_group.add_argument('--pg-password', metavar='PASSWORD', help='PostgreSQL password')
    db_group.add_argument('--pg-schema', metavar='SCHEMA', default='public',
                          help='PostgreSQL schema (default: public)')

    db_group.add_argument('--mysql-host', metavar='HOST', help='MySQL host')
    db_group.add_argument('--mysql-port', metavar='PORT', type=int, default=3306,
                          help='MySQL port (default: 3306)')
    db_group.add_argument('--mysql-db', metavar='DBNAME', help='MySQL database name')
    db_group.add_argument('--mysql-user', metavar='USER', help='MySQL user')
    db_group.add_argument('--mysql-password', metavar='PASSWORD', help='MySQL password')

    db_group.add_argument('--row-limit', metavar='N', type=int, default=10000,
                          help='Max rows to scan per database table (default: 10 000)')

    # ── Cloud storage sources ─────────────────────────────────────────────────
    cloud_group = parser.add_argument_group('Cloud storage sources')
    cloud_group.add_argument('--s3-bucket', metavar='BUCKET',
                             help='Scan an Amazon S3 bucket')
    cloud_group.add_argument('--s3-prefix', metavar='PREFIX', default='',
                             help='S3 key prefix filter (default: all objects)')
    cloud_group.add_argument('--s3-region', metavar='REGION',
                             help='AWS region override')

    cloud_group.add_argument('--gcs-bucket', metavar='BUCKET',
                             help='Scan a Google Cloud Storage bucket')
    cloud_group.add_argument('--gcs-prefix', metavar='PREFIX', default='',
                             help='GCS blob name prefix filter')

    cloud_group.add_argument('--azure-container', metavar='NAME',
                             help='Scan an Azure Blob Storage container')
    cloud_group.add_argument('--azure-prefix', metavar='PREFIX', default='',
                             help='Azure blob name prefix filter')
    cloud_group.add_argument('--azure-conn-string', metavar='STRING',
                             help='Azure Storage connection string '
                                  '(or set AZURE_STORAGE_CONNECTION_STRING)')

    # ── Output ────────────────────────────────────────────────────────────────
    parser.add_argument('--output', metavar='FILE',
                        help='Save findings report to a CSV file')

    args = parser.parse_args()
    detector = CreditCardDetector()
    findings = []

    # ── Dispatch ──────────────────────────────────────────────────────────────

    if args.csv:
        print(f"Scanning CSV: {args.csv}")
        findings = detector.scan_csv(args.csv, args.delimiter)

    elif args.file:
        print(f"Scanning file: {args.file}")
        findings = detector.scan_text_file(args.file)

    elif args.pdf:
        print(f"Scanning PDF: {args.pdf}")
        findings = detector.scan_pdf(args.pdf)

    elif args.excel:
        print(f"Scanning Excel: {args.excel}")
        findings = detector.scan_excel(args.excel)

    elif args.directory:
        print(f"Scanning directory: {args.directory}")
        findings = detector.scan_directory(args.directory)

    elif args.sqlite:
        print(f"Scanning SQLite: {args.sqlite}")
        findings = detector.scan_sqlite(args.sqlite, row_limit=args.row_limit)

    elif args.pg_host:
        if not all([args.pg_db, args.pg_user, args.pg_password]):
            parser.error("PostgreSQL requires --pg-host, --pg-db, --pg-user, --pg-password")
        print(f"Scanning PostgreSQL: {args.pg_host}/{args.pg_db} (schema: {args.pg_schema})")
        findings = detector.scan_postgres(
            args.pg_host, args.pg_db, args.pg_user, args.pg_password,
            port=args.pg_port, schema=args.pg_schema, row_limit=args.row_limit,
        )

    elif args.mysql_host:
        if not all([args.mysql_db, args.mysql_user, args.mysql_password]):
            parser.error("MySQL requires --mysql-host, --mysql-db, --mysql-user, --mysql-password")
        print(f"Scanning MySQL: {args.mysql_host}/{args.mysql_db}")
        findings = detector.scan_mysql(
            args.mysql_host, args.mysql_db, args.mysql_user, args.mysql_password,
            port=args.mysql_port, row_limit=args.row_limit,
        )

    elif args.s3_bucket:
        print(f"Scanning S3: s3://{args.s3_bucket}/{args.s3_prefix}")
        findings = detector.scan_s3(
            args.s3_bucket, prefix=args.s3_prefix, region=args.s3_region,
        )

    elif args.gcs_bucket:
        print(f"Scanning GCS: gs://{args.gcs_bucket}/{args.gcs_prefix}")
        findings = detector.scan_gcs(args.gcs_bucket, prefix=args.gcs_prefix)

    elif args.azure_container:
        print(f"Scanning Azure Blob: {args.azure_container}/{args.azure_prefix}")
        findings = detector.scan_azure_blob(
            args.azure_container, prefix=args.azure_prefix,
            connection_string=args.azure_conn_string,
        )

    else:
        parser.print_help()
        return

    detector.generate_report(findings, args.output)


if __name__ == "__main__":
    main()
